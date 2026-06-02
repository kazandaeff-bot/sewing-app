import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.properties.creator = "Z.ai"

# Style definitions
header_font = Font(name="DejaVu Sans", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
subheader_font = Font(name="DejaVu Sans", bold=True, size=10, color="1B4332")
subheader_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
normal_font = Font(name="DejaVu Sans", size=10)
bold_font = Font(name="DejaVu Sans", bold=True, size=10)
title_font = Font(name="DejaVu Sans", bold=True, size=14, color="1B4332")
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_cell(ws, row, col, bold=False):
    cell = ws.cell(row=row, column=col)
    cell.font = bold_font if bold else normal_font
    cell.alignment = wrap_align
    cell.border = thin_border

def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted

# ==================== SHEET 1: Сервер и доступ ====================
ws1 = wb.active
ws1.title = "1. Сервер и доступ"

ws1.merge_cells('A1:D1')
ws1.cell(row=1, column=1, value="Сервер и доступ к платформе").font = title_font

row = 3
headers = ["Параметр", "Значение", "Статус", "Примечание"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, len(headers))

data = [
    ["URL сервера", "http://0.0.0.0:3000", "✅ Работает", "Production build, next start"],
    ["Авторизация", "POST /api/auth/login", "✅ Работает", "JWT токен, Bearer auth"],
    ["Проверка сессии", "GET /api/auth/me", "✅ Работает", "Возвращает user или null"],
    ["Выход", "POST /api/auth/logout", "✅ Работает", "Инвалидирует сессию"],
    ["Смена пароля", "POST /api/auth/change-password", "✅ Работает", "Минимум 6 символов"],
    ["База данных", "SQLite (prisma/dev.db)", "✅ Работает", "Prisma ORM 6.11.1"],
    ["Runtime", "Node.js + Next.js 16.2.6", "✅ Работает", "Turbopack"],
]

for i, d in enumerate(data):
    r = row + 1 + i
    for j, v in enumerate(d):
        ws1.cell(row=r, column=j+1, value=v)
        style_data_cell(ws1, r, j+1, bold=(j==0))

auto_width(ws1)

# ==================== SHEET 2: Данные для входа ====================
ws2 = wb.create_sheet("2. Данные для входа")

ws2.merge_cells('A1:E1')
ws2.cell(row=1, column=1, value="Данные для входа по ролям").font = title_font

row = 3
headers = ["Роль", "Логин", "Пароль", "ФИО в системе", "Назначение / Доступ"]
for i, h in enumerate(headers, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, len(headers))

users = [
    ["supervisor (Директор)", "admin", "admin", "Директор Ольга", "Полный доступ ко всем разделам. Создание/утверждение планов, управление сотрудниками, справочники, финансы."],
    ["sewer (Швея)", "sewer1", "123456", "Иванова Мария", "Доступ только к разделу «Работа швей». Видит свои задания, отмечает прогресс пошива."],
    ["qc (ОТК)", "qc1", "123456", "Смирнова Ирина", "Доступ к разделу «ОТК». Проверка качества, приёмка/возврат на переделку."],
    ["seller (Продавец)", "seller1", "123456", "Продавец Анна", "Перенаправляется в Производство. Раздел в разработке."],
    ["technologist (Технолог)", "technologist1", "123456", "Технолог Мария", "Заглушка — «Раздел в разработке»."],
    ["cutter (Закройщик)", "cutter1", "123456", "Закройщик Иван", "Заглушка — «Раздел в разработке»."],
    ["ironing (Утюжильщик)", "ironing1", "123456", "Норгиза", "Доступ к разделу «ВТО». Отмечает позиции как отглаженные."],
    ["customer (Заказчик)", "textilpro", "123456", "ООО ТекстильПро (кабинет)", "Просмотр своих планов пошива, распределение по городам, упаковка, материалы."],
]

for i, u in enumerate(users):
    r = row + 1 + i
    for j, v in enumerate(u):
        ws2.cell(row=r, column=j+1, value=v)
        style_data_cell(ws2, r, j+1, bold=(j==0))

auto_width(ws2)
ws2.column_dimensions['E'].width = 60

# ==================== SHEET 3: Бизнес-процессы ====================
ws3 = wb.create_sheet("3. Бизнес-процессы")

ws3.merge_cells('A1:C1')
ws3.cell(row=1, column=1, value="Бизнес-процессы платформы").font = title_font

processes = [
    ["1. Приём заказа", 
     "1. Супервайзер создаёт «План пошива» (раздел Планы пошива)\n"
     "   — Быстрое создание: пустой черновик, затем добавление позиций через редактирование\n"
     "   — Полное создание: выбор заказчика, приоритет, дедлайн, позиции (изделие+размер+цвет+кол-во)\n"
     "2. План получает статус «Черновик» (draft)\n"
     "3. Супервайзер утверждает план → статус «Утверждён» (approved)\n"
     "4. При утверждении автоматически создаётся План раскроя (cutting plan) с позициями\n"
     "5. Статус: draft → approved → in_work → shipped → shipped_paid",
     "Супервайзер (создание, утверждение), система (авто-создание раскроя)"],
    
    ["2. Раскрой",
     "1. При утверждении плана автоматически создаётся CuttingPlan со статусом «В работе» (in_work)\n"
     "2. Закройщик/супервайзер заполняет фактическое кол-во (actualQty) и кол-во пачек (bundleCount)\n"
     "3. Элементы сгруппированы по цвету, размеры отсортированы от меньшего к большему\n"
     "4. При завершении раскроя — статус «Раскроено» (cut)\n"
     "5. Если фактическое кол-во > планового → создаётся остаток кроя (CuttingLeftover)\n"
     "6. Возможна печать документа раскроя",
     "Супервайзер (управление), система (авто-создание)"],
    
    ["3. Распределение задач по швеям",
     "1. Из раскроя (CuttingPlan) супервайзер создаёт задания швеям (SewingTask)\n"
     "   — Выбор раскроя, выбор швеи, добавление позиций (изделие+размер+цвет+кол-во)\n"
     "2. Задание получает статус «Выдано» (issued)\n"
     "3. Статусы: issued → in_work → pending_ironing → pending_qc → completed",
     "Супервайзер (назначение), швея (выполнение)"],
    
    ["4. Пошив",
     "1. Швея видит свои активные задания в разделе «Работа швей»\n"
     "2. Берёт задание в работу → статус «В работе» (in_work)\n"
     "3. Отмечает фактическое кол-во по каждой позиции\n"
     "4. Фиксирует брак ткани (fabricDefect) и примечание\n"
     "5. Завершает пошив → статус «Ожидает ВТО» (pending_ironing)",
     "Швея (выполнение пошива)"],
    
    ["5. Контроль качества (ОТК)",
     "1. Сотрудник ОТК видит задания со статусом pending_qc\n"
     "2. Проверяет качество, подтверждает кол-во или возвращает на переделку\n"
     "3. При возврате создаётся запись SewingRework с причиной и кол-вом\n"
     "4. Переделка проходит: pending → in_progress → pending_qc → completed\n"
     "5. После проверки → статус «Завершено» (completed)\n"
     "6. При отгрузке плана начисляется зарплата ОТК (qcSalary)",
     "Сотрудник ОТК (проверка), система (начисление зарплаты)"],
    
    ["6. Глажка (ВТО)",
     "1. Утюжильщик видит задания со статусом pending_ironing\n"
     "2. Отмечает позиции как отглаженные\n"
     "3. После глажки → статус «Ожидает ОТК» (pending_qc)",
     "Утюжильщик (выполнение ВТО)"],
    
    ["7. Отгрузка",
     "1. Когда план в статусе in_work и все задачи завершены\n"
     "2. Супервайзер выбирает сотрудника ОТК для начисления зарплаты\n"
     "3. Переводит план в статус «Отгружен» (shipped)\n"
     "4. Автоматически создаётся запись qcSalary с расчётом зарплаты ОТК\n"
     "5. После оплаты → статус «Отгружен и оплачен» (shipped_paid)",
     "Супервайзер (отгрузка, оплата), система (расчёт зарплаты)"],
]

row = 3
headers = ["Процесс", "Шаги", "Участники"]
for i, h in enumerate(headers, 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, len(headers))

for i, p in enumerate(processes):
    r = row + 1 + i
    for j, v in enumerate(p):
        ws3.cell(row=r, column=j+1, value=v)
        style_data_cell(ws3, r, j+1, bold=(j==0))

auto_width(ws3)
ws3.column_dimensions['B'].width = 80
ws3.column_dimensions['C'].width = 40

# ==================== SHEET 4: Готовность модулей ====================
ws4 = wb.create_sheet("4. Готовность модулей")

ws4.merge_cells('A1:D1')
ws4.cell(row=1, column=1, value="Готовность модулей").font = title_font

row = 3
headers = ["Модуль / Страница", "Раздел в меню", "Статус", "Примечание"]
for i, h in enumerate(headers, 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, len(headers))

modules = [
    ["Планы пошива (SewingPlansTab)", "Планирование", "✅ Готово", "Создание, редактирование, утверждение, дополнение, удаление. Быстрое создание. Детальный просмотр. Фильтр по заказчику. Приоритеты и дедлайны."],
    ["Лекала (PatternsTab)", "Производство", "✅ Готово", "Управление лекалами изделий. Привязка к продуктам."],
    ["Раскладка (NestingTab)", "Производство", "✅ Готово", "Визуальная раскладка лекал. Оптимизация расхода ткани."],
    ["Раскрой (CuttingPlansTab)", "Производство", "✅ Готово", "Просмотр раскроев. Группировка по цветам, сортировка размеров. Факт, пачки, печать."],
    ["Остатки раскроя (CuttingLeftoversTab)", "Производство", "✅ Готово", "Просмотр и учёт остатков кроя. Статусы: свободно, частично пошито, полностью пошито."],
    ["Задания швеям (SewingTasksTab)", "Производство", "✅ Готово", "Создание заданий из раскроя. Назначение швеи. Статусы заданий."],
    ["Работа швей (SewerTab)", "Производство", "✅ Готово", "Швея видит свои задания. Отмечает прогресс, брак. Фильтр по швее."],
    ["ОТК (QCTab)", "Производство", "✅ Готово", "Проверка качества. Приёмка, возврат на переделку. Учёт переделок."],
    ["ВТО / Глажка (IroningTab)", "Производство", "✅ Готово", "Отметка отглаженных позиций."],
    ["Упаковка (BoxesTab)", "Производство", "✅ Готово", "Автоматическое распределение по коробам. Формирование, проверка, отгрузка."],
    ["Города / Распределение (CityDistributionTab)", "Планирование", "✅ Готово", "Планы продавцов с распределением по городам. Создание, утверждение, распределение."],
    ["Сделки / CRM (CRMTab)", "Финансы", "✅ Готово", "Воронка сделок. Переговоры, контакты, статусы."],
    ["Договоры (ContractsTab)", "Финансы", "✅ Готово", "CRUD договоров. Типы: услуга / поставка. Статусы, реквизиты."],
    ["Счета (InvoicesTab)", "Финансы", "✅ Готово", "Выставление счетов. Позиции, НДС, привязка к плану."],
    ["УПД (UPDTab)", "Финансы", "✅ Готово", "Универсальные передаточные документы. Привязка к счетам и отгрузкам."],
    ["Зарплата ОТК (QcSalariesTab)", "Финансы", "✅ Готово", "Автоматический расчёт зарплаты ОТК при отгрузке. Просмотр, отметка оплаты."],
    ["Изделия (ProductsTab)", "Справочники", "✅ Готово", "CRUD изделий. Артикул, ставки, размеры, цвета, комплекты. Ставки по размерам."],
    ["Сотрудники (EmployeesTab)", "Справочники", "✅ Готово", "CRUD сотрудников. Роли, коды, логины. Сброс пароля."],
    ["Прочее / Справочники (ReferencesTab)", "Справочники", "✅ Готово", "Типы материалов, материалы, нормы расхода, города, типы коробов, причины переделки."],
    ["Материалы заказчика (CustomerMaterialsTab)", "Заказчик", "✅ Готово", "Просмотр баланса материалов. Доступно только для роли customer."],
    ["Продавец (seller role)", "—", "🔧 Заглушка", "Перенаправление в Производство. Полноценный раздел не создан."],
    ["Технолог (technologist role)", "—", "🔧 Заглушка", "Отображается «Раздел в разработке». Нет доступа к модулям."],
    ["Закройщик (cutter role)", "—", "🔧 Заглушка", "Отображается «Раздел в разработке». Нет доступа к модулям."],
]

green_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

for i, m in enumerate(modules):
    r = row + 1 + i
    for j, v in enumerate(m):
        ws4.cell(row=r, column=j+1, value=v)
        style_data_cell(ws4, r, j+1, bold=(j==0))
    # Color by status
    status = m[2]
    if "✅" in status:
        ws4.cell(row=r, column=3).fill = green_fill
    elif "🔧" in status:
        ws4.cell(row=r, column=3).fill = yellow_fill
    elif "❌" in status or "🐛" in status:
        ws4.cell(row=r, column=3).fill = red_fill

auto_width(ws4)
ws4.column_dimensions['D'].width = 70

# ==================== SHEET 5: Известные проблемы ====================
ws5 = wb.create_sheet("5. Проблемы и TODO")

ws5.merge_cells('A1:C1')
ws5.cell(row=1, column=1, value="Известные проблемы и TODO").font = title_font

row = 3
headers = ["Категория", "Описание", "Приоритет"]
for i, h in enumerate(headers, 1):
    ws5.cell(row=row, column=i, value=h)
style_header_row(ws5, row, len(headers))

issues = [
    ["Баг", "Роли seller, technologist, cutter — полноценные разделы не реализованы, отображаются заглушки", "Средний"],
    ["Баг", "Production-сервер (standalone) крашится после нескольких запросов — используется next start вместо standalone", "Высокий"],
    ["Баг", "Sidebar не включает «Лекала» и «Раскладка» для роли supervisor (они доступны через URL/код, но не в меню)", "Средний"],
    ["Ограничение", "Seed endpoint заблокирован в production (NODE_ENV=production) — нельзя пересоздать тестовые данные", "Низкий"],
    ["Ограничение", "Нет пагинации на списках — при большом количестве записей возможны задержки", "Средний"],
    ["Ограничение", "Материалы — нет полноценного модуля учёта поступления/списания в UI (только через API)", "Средний"],
    ["TODO", "Полноценный раздел для закройщика (просмотр раскроев, отметка факта)", "Средний"],
    ["TODO", "Полноценный раздел для технолога (управление лекалами, тех. карты)", "Средний"],
    ["TODO", "Полноценный раздел для продавца (планы продаж, отгрузки)", "Средний"],
    ["TODO", "Сводка/дашборд на главной странице с ключевыми метриками", "Низкий"],
    ["TODO", "Уведомления (push/email) при смене статусов", "Низкий"],
    ["TODO", "Экспорт отчётов (PDF/Excel) по планам, зарплате, отгрузкам", "Средний"],
    ["TODO", "Интеграция с 1С для обмена документами (УПД, счета)", "Низкий"],
]

for i, issue in enumerate(issues):
    r = row + 1 + i
    for j, v in enumerate(issue):
        ws5.cell(row=r, column=j+1, value=v)
        style_data_cell(ws5, r, j+1, bold=(j==0))

auto_width(ws5)
ws5.column_dimensions['B'].width = 80

# ==================== SHEET 6: API эндпоинты ====================
ws6 = wb.create_sheet("6. API эндпоинты")

ws6.merge_cells('A1:D1')
ws6.cell(row=1, column=1, value="Список всех API эндпоинтов").font = title_font

row = 3
headers = ["Метод", "Путь", "Описание", "Доступ"]
for i, h in enumerate(headers, 1):
    ws6.cell(row=row, column=i, value=h)
style_header_row(ws6, row, len(headers))

endpoints = [
    # Auth
    ["POST", "/api/auth/login", "Авторизация (возвращает JWT токен + user)", "Открытый"],
    ["POST", "/api/auth/logout", "Выход из системы", "Авторизованный"],
    ["GET", "/api/auth/me", "Проверка текущей сессии", "Авторизованный"],
    ["POST", "/api/auth/change-password", "Смена пароля", "Авторизованный"],
    ["GET", "/api/auth/session", "Информация о сессии", "Авторизованный"],
    # Employees
    ["GET", "/api/employees", "Список сотрудников", "supervisor"],
    ["POST", "/api/employees", "Создать сотрудника", "supervisor"],
    ["PATCH", "/api/employees/[id]", "Обновить сотрудника", "supervisor"],
    ["DELETE", "/api/employees/[id]", "Удалить сотрудника", "supervisor"],
    ["POST", "/api/employees/[id]/reset-password", "Сброс пароля сотрудника", "supervisor"],
    # Products
    ["GET", "/api/products", "Список изделий", "supervisor, customer"],
    ["POST", "/api/products", "Создать изделие", "supervisor"],
    ["PATCH", "/api/products/[id]", "Обновить изделие", "supervisor"],
    ["DELETE", "/api/products/[id]", "Удалить изделие", "supervisor"],
    # Product Size Rates
    ["GET", "/api/product-size-rates?productId=...", "Ставки по размерам изделия", "supervisor"],
    ["POST", "/api/product-size-rates", "Создать ставку по размеру", "supervisor"],
    ["PATCH", "/api/product-size-rates", "Обновить ставку", "supervisor"],
    ["DELETE", "/api/product-size-rates", "Удалить ставку", "supervisor"],
    # Customers
    ["GET", "/api/customers", "Список заказчиков", "supervisor"],
    ["POST", "/api/customers", "Создать заказчика", "supervisor"],
    ["PATCH", "/api/customers/[id]", "Обновить заказчика", "supervisor"],
    ["DELETE", "/api/customers/[id]", "Удалить заказчика", "supervisor"],
    # Plans
    ["GET", "/api/plans", "Список планов пошива", "supervisor, customer"],
    ["POST", "/api/plans", "Создать план пошива", "supervisor"],
    ["GET", "/api/plans/[id]", "Детали плана с прогрессом", "supervisor"],
    ["PATCH", "/api/plans/[id]", "Обновить план (статус, позиции)", "supervisor"],
    ["DELETE", "/api/plans/[id]", "Удалить план", "supervisor"],
    # Cutting Plans
    ["GET", "/api/cutting-plans", "Список планов раскроя", "supervisor"],
    ["PATCH", "/api/cutting-plans/[id]", "Обновить раскрой (статус, факт, пачки)", "supervisor"],
    # Cutting Leftovers
    ["GET", "/api/cutting-leftovers", "Список остатков кроя", "supervisor"],
    ["POST", "/api/cutting-leftovers", "Создать остаток кроя", "supervisor"],
    ["PATCH", "/api/cutting-leftovers/[id]", "Обновить остаток", "supervisor"],
    ["DELETE", "/api/cutting-leftovers/[id]", "Удалить остаток", "supervisor"],
    # Sewing Tasks
    ["GET", "/api/sewing-tasks", "Список заданий швеям", "supervisor"],
    ["POST", "/api/sewing-tasks", "Создать задание швее", "supervisor"],
    ["PATCH", "/api/sewing-tasks/[id]", "Обновить задание (статус, позиции)", "supervisor, sewer"],
    ["DELETE", "/api/sewing-tasks/[id]", "Удалить задание", "supervisor"],
    # Sewing Reworks
    ["GET", "/api/sewing-reworks", "Список переделок", "supervisor"],
    ["POST", "/api/sewing-reworks", "Создать переделку", "supervisor, qc"],
    ["PATCH", "/api/sewing-reworks/[id]", "Обновить статус переделки", "supervisor, qc"],
    # Tasks (legacy)
    ["GET", "/api/tasks", "Список задач (старая система)", "supervisor"],
    ["POST", "/api/tasks", "Создать задачу", "supervisor"],
    ["PATCH", "/api/tasks/[id]", "Обновить задачу", "supervisor, sewer"],
    ["DELETE", "/api/tasks/[id]", "Удалить задачу", "supervisor"],
    # Ironing
    ["GET", "/api/ironing", "Список позиций для ВТО", "supervisor, ironing"],
    ["PATCH", "/api/ironing", "Отметить позиции как отглаженные", "supervisor, ironing"],
    # Stats
    ["GET", "/api/stats", "Статистика дашборда", "supervisor"],
    # Seller Plans
    ["GET", "/api/seller-plans", "Список планов продавцов", "supervisor"],
    ["POST", "/api/seller-plans", "Создать план продавца", "supervisor"],
    ["GET", "/api/seller-plans/[id]", "Детали плана продавца", "supervisor"],
    ["PATCH", "/api/seller-plans/[id]", "Обновить план продавца", "supervisor"],
    ["DELETE", "/api/seller-plans/[id]", "Удалить план продавца", "supervisor"],
    ["GET", "/api/seller-plans/available-items?customerId=...", "Доступные позиции для плана", "supervisor"],
    # Boxes
    ["GET", "/api/boxes", "Список коробов", "supervisor"],
    ["POST", "/api/boxes", "Создать короб", "supervisor"],
    ["PATCH", "/api/boxes/[id]", "Обновить короб", "supervisor"],
    ["GET", "/api/boxes/[id]/print", "Печать упаковочного листа", "supervisor"],
    # Box Types
    ["GET", "/api/box-types", "Список типов коробов", "supervisor"],
    ["POST", "/api/box-types", "Создать тип короба", "supervisor"],
    ["PATCH", "/api/box-types/[id]", "Обновить тип короба", "supervisor"],
    ["DELETE", "/api/box-types/[id]", "Удалить тип короба", "supervisor"],
    # Cities
    ["GET", "/api/cities", "Список городов", "supervisor"],
    ["POST", "/api/cities", "Создать город", "supervisor"],
    ["DELETE", "/api/cities/[id]", "Удалить город", "supervisor"],
    # Deals (CRM)
    ["GET", "/api/deals", "Список сделок", "supervisor"],
    ["POST", "/api/deals", "Создать сделку", "supervisor"],
    ["GET", "/api/deals/[id]", "Детали сделки", "supervisor"],
    ["PATCH", "/api/deals/[id]", "Обновить сделку", "supervisor"],
    ["DELETE", "/api/deals/[id]", "Удалить сделку", "supervisor"],
    ["GET", "/api/deals/[id]/contacts", "Контакты сделки", "supervisor"],
    ["POST", "/api/deals/[id]/contacts", "Добавить контакт", "supervisor"],
    ["PATCH", "/api/deals/[id]/contacts/[contactId]", "Обновить контакт", "supervisor"],
    ["DELETE", "/api/deals/[id]/contacts/[contactId]", "Удалить контакт", "supervisor"],
    # Contracts
    ["GET", "/api/contracts", "Список договоров", "supervisor"],
    ["POST", "/api/contracts", "Создать договор", "supervisor"],
    ["GET", "/api/contracts/[id]", "Детали договора", "supervisor"],
    ["PATCH", "/api/contracts/[id]", "Обновить договор", "supervisor"],
    ["DELETE", "/api/contracts/[id]", "Удалить договор", "supervisor"],
    # Invoices
    ["GET", "/api/invoices", "Список счетов", "supervisor"],
    ["POST", "/api/invoices", "Создать счёт", "supervisor"],
    ["GET", "/api/invoices/[id]", "Детали счёта", "supervisor"],
    ["PATCH", "/api/invoices/[id]", "Обновить счёт", "supervisor"],
    ["DELETE", "/api/invoices/[id]", "Удалить счёт", "supervisor"],
    # UPD
    ["GET", "/api/upd", "Список УПД", "supervisor"],
    ["POST", "/api/upd", "Создать УПД", "supervisor"],
    ["GET", "/api/upd/[id]", "Детали УПД", "supervisor"],
    ["PATCH", "/api/upd/[id]", "Обновить УПД", "supervisor"],
    ["DELETE", "/api/upd/[id]", "Удалить УПД", "supervisor"],
    # Materials
    ["GET", "/api/material-types", "Список типов материалов", "supervisor"],
    ["POST", "/api/material-types", "Создать тип материала", "supervisor"],
    ["PATCH", "/api/material-types/[id]", "Обновить тип материала", "supervisor"],
    ["DELETE", "/api/material-types/[id]", "Удалить тип материала", "supervisor"],
    ["GET", "/api/materials", "Список материалов", "supervisor, customer"],
    ["POST", "/api/materials", "Создать материал", "supervisor"],
    ["PATCH", "/api/materials/[id]", "Обновить материал", "supervisor"],
    ["DELETE", "/api/materials/[id]", "Удалить материал", "supervisor"],
    ["GET", "/api/material-entries", "Движение материалов", "supervisor, customer"],
    ["POST", "/api/material-entries", "Создать запись движения", "supervisor"],
    ["DELETE", "/api/material-entries/[id]", "Удалить запись", "supervisor"],
    ["GET", "/api/material-norms", "Нормы расхода материалов", "supervisor"],
    ["POST", "/api/material-norms", "Создать норму расхода", "supervisor"],
    ["PATCH", "/api/material-norms/[id]", "Обновить норму", "supervisor"],
    ["DELETE", "/api/material-norms/[id]", "Удалить норму", "supervisor"],
    ["GET", "/api/material-balance?customerId=...", "Баланс материалов заказчика", "supervisor, customer"],
    # Print
    ["GET", "/api/print?type=...&id=...", "Печать документа (раскрой, задание, упаковочный лист)", "supervisor"],
    # Seed
    ["GET", "/api/seed", "Инициализация тестовых данных (только dev)", "supervisor"],
]

# Color for HTTP methods
method_colors = {
    "GET": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "POST": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "PATCH": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "DELETE": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
}

for i, ep in enumerate(endpoints):
    r = row + 1 + i
    for j, v in enumerate(ep):
        ws6.cell(row=r, column=j+1, value=v)
        style_data_cell(ws6, r, j+1)
    # Color method cell
    method = ep[0]
    if method in method_colors:
        ws6.cell(row=r, column=1).fill = method_colors[method]
        ws6.cell(row=r, column=1).alignment = center_align

auto_width(ws6)
ws6.column_dimensions['B'].width = 45
ws6.column_dimensions['C'].width = 55
ws6.column_dimensions['D'].width = 25

# Save
output_path = "/home/z/my-project/download/audit_platform.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
